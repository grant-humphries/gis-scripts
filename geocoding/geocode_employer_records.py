import argparse
import csv
import requests
import sys
from collections import defaultdict, OrderedDict
from functools import partial
from os.path import abspath, dirname, join

import fiona
import pyproj
from openpyxl import load_workbook
from rtree import index
from shapely import ops
from shapely.geometry import mapping, shape, Point

HOME = dirname(dirname(abspath(__file__)))
EMP_FILE_NAME = 'TriMet Employer List 2015 - Oct 15.xlsx'
EMPLOYERS = join(HOME, 'employer_list', EMP_FILE_NAME)
GEO_FILE_NAME = '{} Geocoded.xlsx'.format(EMP_FILE_NAME.split('.')[0])
GEOCODED = join(HOME, 'xlsx', GEO_FILE_NAME)
UNGEOCODEABLE = join(HOME, 'csv', 'ungeocodeable_addresses.csv')
RAIL_STOP = '//gisstore/gis/TRIMET/rail_stop.shp'
EMP_STATIONS = join(HOME, 'csv', 'employers_half_mile_max_orange_v2.csv')


def get_ospn_coordinates_for_employers():
    """"""

    rlis_token = process_options().rlis_token
    wgs2ospn = create_transformation('4326', '2913')
    ungeocodeable = list()

    wb = load_workbook(EMPLOYERS)
    ws = wb.worksheets[0]

    # add headers for the columns that will hold the x and y coordinates
    for new_field in ('X', 'Y'):
        cell_value = '{} Coordinate'.format(new_field)
        ws.cell(row=1, column=ws.max_column+1).value = cell_value

    header = [str(cell.value) for cell in ws.rows[0]]
    lat_ix = header.index('Latitude')
    lon_ix = header.index('Longitude')
    addr_ix = header.index('Street Address')
    city_ix = header.index('City')
    state_ix = header.index('State')
    zip_ix = header.index('Zip')
    x_ix = header.index('X Coordinate')
    y_ix = header.index('Y Coordinate')

    for row in ws.iter_rows(row_offset=1):
        lat = row[lat_ix].value
        lon = row[lon_ix].value

        if lat:
            lat, lon = float(lat), float(lon)
            x, y = convert_coordinates(lon, lat, wgs2ospn)
        else:
            # within the spreadsheet some zip values are stored as
            # float which leaves a trailing zero
            zip_code = row[zip_ix].value
            if isinstance(zip_code, float):
                zip_code = int(zip_code)

            addr_str = '{addr}, {city}, {state} {zip}'.format(
                addr=row[addr_ix].value,
                city=row[city_ix].value,
                state=row[state_ix].value,
                zip=zip_code)

            rlis_gc = rlis_geocode(addr_str, rlis_token)
            if rlis_gc:
                x, y = rlis_gc['ORSP_x'], rlis_gc['ORSP_y']
            else:
                google_gc = google_geocode(addr_str)
                if google_gc:
                    lon, lat = google_gc['lng'], google_gc['lat']
                    x, y = convert_coordinates(lon, lat, wgs2ospn)
                else:
                    ungeocodeable.append([cell.value for cell in row])
                    continue

        row[x_ix].value = x
        row[y_ix].value = y

    wb.save(GEOCODED)

    with open(UNGEOCODEABLE, 'wb') as ungeo_csv:
        ungeo_writer = csv.writer(ungeo_csv)
        ungeo_writer.writerow(header[:-2])
        for row in ungeocodeable:
            ungeo_writer.writerow(row)


def process_options():
    """"""

    arglist = sys.argv[1:]

    parser = argparse.ArgumentParser()
    parser.add_argument(
        '-t', '--rlis_token',
        dest='rlis_token',
        required=True,
        help='Token for RLIS API'
    )

    options = parser.parse_args(arglist)
    return options


def create_transformation(src_srs, dst_srs):
    """"""

    transformation = partial(
            pyproj.transform,
            pyproj.Proj(init='epsg:{}'.format(src_srs)),
            pyproj.Proj(init='epsg:{}'.format(dst_srs), preserve_units=True))

    return transformation


def convert_coordinates(x, y, transformation):
    """"""

    src_pt = Point(x, y)
    dst_pt = ops.transform(transformation, src_pt)

    return dst_pt.x, dst_pt.y


def rlis_geocode(addr_str, token):
    """Take an input address string, send it to the rlis api and return
    a dictionary that are the state plane coordinated for that address,
    handle errors in the request fails in one way or another"""

    url = 'http://gis.oregonmetro.gov/rlisapi2/locate/'
    params = {
        'token': token,
        'input': addr_str,
        'form': 'json'
    }
    rsp = requests.get(url, params=params)

    if rsp.status_code != 200:
        print 'unable to establish connection with rlis api'
        print 'status code is: {0}'.format(rsp.status_code)
        return None

    json_rsp = rsp.json()
    if json_rsp['error']:
        print 'the following address could not be geocoded by the rlis api:'
        print "'{0}'".format(addr_str)
        print 'the following error message was returned:'
        print "'{0}'".format(json_rsp['error']), '\n'
        return None
    else:
        return json_rsp['data'][0]


def google_geocode(addr_str):
    """"""

    url = 'https://maps.googleapis.com/maps/api/geocode/json'
    params = {'address': addr_str}
    rsp = requests.get(url, params=params)

    if rsp.status_code != 200:
        print 'unable to establish connection with google geocoder api'
        print 'status code is: {0}'.format(rsp.status_code)
        return None

    json_rsp = rsp.json()
    if json_rsp['status'] == 'OK':
        location = json_rsp['results'][0]['geometry']['location']
        return location
    if json_rsp['status'] == 'OVER_QUERY_LIMIT':
        print 'daily limit exceeded, no more google geocoding today :|'
        return None
    else:
        print 'google geocode was not successful with status code:'
        print json_rsp['status']
        return None


def get_employers_near_stops():
    """"""

    distance = 5280 / 2
    filter_field = 'LINE'
    filter_vals = ['O']

    stop_buffers = dict()
    stop_names = dict()
    with fiona.open(RAIL_STOP) as rail_stop:
        for fid, feat in rail_stop.items():
            fields = feat['properties']
            if fields[filter_field] in filter_vals:
                geom = shape(feat['geometry'])
                feat['geometry'] = mapping(geom.buffer(distance))
                stop_buffers[fid] = feat
                stop_names[fid] = fields['STATION']

    wb = load_workbook(GEOCODED)
    ws = wb.worksheets[0]

    header = [str(cell.value) for cell in ws.rows[0]]
    x_ix = header.index('X Coordinate')
    y_ix = header.index('Y Coordinate')

    employers = dict()
    for i, row in enumerate(ws.iter_rows(row_offset=1)):
        x, y = row[x_ix].value, row[y_ix].value
        if x and y:
            feat = dict()
            x, y = float(x), float(y)
            feat['geometry'] = mapping(Point(x, y))
            field_vals = [cell.value for cell in row]
            feat['properties'] = OrderedDict(zip(header, field_vals))
            employers[i] = feat

    join_mapping = spatial_join(stop_buffers, employers)

    with open(EMP_STATIONS, 'wb') as emp_stations:
        emp_writer = csv.writer(emp_stations)
        station_header = ['Stations'] + header
        emp_writer.writerow(station_header)

        for i, row in enumerate(ws.iter_rows(row_offset=1)):
            if i in join_mapping:
                station_ids = join_mapping[i]
                station_str = ', '.join([stop_names[sid] for sid in station_ids])
                csv_row = [station_str]
                for cell in row:
                    value = cell.value
                    if isinstance(value, unicode):
                        value = value.encode('utf-8')
                    csv_row.append(value)

                emp_writer.writerow(csv_row)


def spatial_join(target_feats, join_feats):
    """this function expects features to be inputting in the format
    generated by fiona, a dictionary with fid as key and a json oject
    as the value with the geometry and attributes"""

    s_index = index.Index()
    join_mapping = defaultdict(list)

    for t_fid, t_feat in target_feats.items():
        t_geom = shape(t_feat['geometry'])
        s_index.insert(t_fid, t_geom.bounds)

    for j_fid, j_feat in join_feats.items():
        j_geom = shape(j_feat['geometry'])

        for t_fid in s_index.intersection(j_geom.bounds):
            t_geom = shape(target_feats[t_fid]['geometry'])
            if j_geom.intersects(t_geom):
                join_mapping[j_fid].append(t_fid)

    return join_mapping


def main():
    """"""

    # get_ospn_coordinates_for_employers()
    get_employers_near_stops()


if __name__ == '__main__':
    main()

